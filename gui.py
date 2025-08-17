import os
import sys
import math
import shutil
import threading
import time
import webbrowser
from datetime import datetime
from tkinter import (
    Tk, StringVar, Text, END, DISABLED, NORMAL, BOTH, RIGHT, LEFT, X, Y,
    filedialog, messagebox, BooleanVar, Canvas
)
from tkinter import ttk
import tkinter.font as tkfont

# Try to import openpyxl (required for .xlsx logs)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None

# ----------------------------
# Utility functions
# ----------------------------
def human_time(ts):
    try:
        return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))
    except Exception:
        return str(ts)

def file_meta(path):
    """Return (name, ext, size, mtime_int, ctime_int). ext is lower without leading dot."""
    name = os.path.basename(path)
    root, ext = os.path.splitext(name)
    ext = ext[1:].lower()
    try:
        size = os.path.getsize(path)
    except Exception:
        size = -1
    try:
        mtime = int(os.path.getmtime(path))
    except Exception:
        mtime = -1
    try:
        ctime = int(os.path.getctime(path))  # On Linux this is metadata-change time
    except Exception:
        ctime = -1
    return name, ext, size, mtime, ctime

def files_identical(src_path, dst_path):
    """Compare metadata (filename, filetype(ext), size, mtime)."""
    if not os.path.exists(dst_path):
        return False
    s_name, s_ext, s_size, s_mtime, _ = file_meta(src_path)
    d_name, d_ext, d_size, d_mtime, _ = file_meta(dst_path)
    return (s_name == d_name and
            s_ext == d_ext and
            s_size == d_size and
            s_mtime == d_mtime)

def next_available_name(dst_dir, filename):
    """Return a non-colliding filename by adding -1, -2, ... before extension."""
    base, ext = os.path.splitext(filename)
    candidate = filename
    counter = 1
    while os.path.exists(os.path.join(dst_dir, candidate)):
        candidate = f"{base}-{counter}{ext}"
        counter += 1
    return candidate

# ----------------------------
# Main App
# ----------------------------
class MoveApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Ragilmalik's Python GUI Mover")
        self.root.geometry("980x660")
        self.root.minsize(860, 580)

        # Vars
        self.src_var = StringVar()
        self.dst_var = StringVar()
        self.status_var = StringVar(value="Ready")
        self.dry_run_var = BooleanVar(value=True)   # Simulation Only default ON
        self.theme_var = StringVar(value="Dark")    # default theme
        self.csv_loc_var = StringVar(value="dest")  # dest | source | custom
        self.csv_custom_dir = StringVar(value="")   # custom dir path
        self._worker_thread = None
        self._last_log_path = None  # .xlsx path

        # Styling / theme
        self._setup_style_palettes()
        self._apply_theme(self.theme_var.get())
        self._apply_picker_highlights(self.theme_var.get())

        # Header (modern gradient)
        self._build_header()

        # Body controls
        self._build_controls()

        # Progress + log
        self._build_progress()
        self._build_log()

    # ---------- Theme & Styles ----------
    def _setup_style_palettes(self):
        # Pure black / pure white bases, opposite text colors
        self.palettes = {
            "Dark": {
                "bg": "#000000",           # pure black
                "panel": "#0a0a0a",
                "card": "#111111",
                "text": "#ffffff",         # opposite: white text
                "muted": "#c9c9c9",
                "accent": "#6d28d9",       # violet-700 (Run)
                "accent_hover": "#7c3aed",
                "btn_clear": "#2563eb",    # blue
                "btn_clear_hover": "#3b82f6",
                "btn_delete": "#dc2626",   # red
                "btn_delete_hover": "#ef4444",
                "btn_open": "#059669",     # green
                "btn_open_hover": "#10b981",
                "btn_exit": "#4b5563",     # gray
                "btn_exit_hover": "#6b7280",
                "entry_bg": "#0b0b0b",
                "entry_focus": "#1a1a1a",
                "progress_trough": "#0b0b0b",
                "gradient_from": "#222222",
                "gradient_to": "#111111",
                "text_bg": "#0b0b0b",
                "text_fg": "#ffffff",
                "text_ins": "#ffffff",
                "picker_highlight_bg": "#ffffff",  # opposite of dark
                "picker_highlight_fg": "#000000",
            },
            "Light": {
                "bg": "#ffffff",           # pure white
                "panel": "#f2f2f2",
                "card": "#ffffff",
                "text": "#000000",         # opposite: black text
                "muted": "#555555",
                "accent": "#7c3aed",
                "accent_hover": "#6d28d9",
                "btn_clear": "#2563eb",
                "btn_clear_hover": "#1d4ed8",
                "btn_delete": "#dc2626",
                "btn_delete_hover": "#b91c1c",
                "btn_open": "#059669",
                "btn_open_hover": "#047857",
                "btn_exit": "#6b7280",
                "btn_exit_hover": "#4b5563",
                "entry_bg": "#ffffff",
                "entry_focus": "#e5e5e5",
                "progress_trough": "#e5e5e5",
                "gradient_from": "#eaeaea",
                "gradient_to": "#f8f8f8",
                "text_bg": "#ffffff",
                "text_fg": "#000000",
                "text_ins": "#000000",
                "picker_highlight_bg": "#000000",  # opposite of light
                "picker_highlight_fg": "#ffffff",
            }
        }

        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except Exception:
            pass

        # Fonts
        self.font_title = tkfont.Font(size=16, weight="bold")
        self.font_subtle = tkfont.Font(size=10)
        self.font_ui = tkfont.Font(size=10)
        self.font_button = tkfont.Font(size=10, weight="bold")

    def _apply_theme(self, mode: str):
        p = self.palettes[mode]

        # Root BG
        self.root.configure(bg=p["bg"])

        # Global
        self.style.configure(".", background=p["bg"], foreground=p["text"], fieldbackground=p["panel"])

        # Frames
        self.style.configure("Card.TFrame", background=p["card"])
        self.style.configure("Panel.TFrame", background=p["panel"])

        # Labels
        self.style.configure("TLabel", background=p["bg"], foreground=p["text"])
        self.style.configure("Muted.TLabel", background=p["bg"], foreground=p["muted"])

        # Entry
        self.style.configure("TEntry", fieldbackground=p["entry_bg"], foreground=p["text"], insertcolor=p["text"], padding=6)
        self.style.map("TEntry", fieldbackground=[("focus", p["entry_focus"])])

        # Buttons — distinct colors
        self.style.configure("Run.TButton", padding=10, relief="flat", background=p["accent"], foreground="white")
        self.style.map("Run.TButton", background=[("active", p["accent_hover"])], relief=[("pressed","flat")])

        self.style.configure("Clear.TButton", padding=10, relief="flat", background=p["btn_clear"], foreground="white")
        self.style.map("Clear.TButton", background=[("active", p["btn_clear_hover"])])

        we = p  # alias
        self.style.configure("DeleteCsv.TButton", padding=10, relief="flat", background=we["btn_delete"], foreground="white")
        self.style.map("DeleteCsv.TButton", background=[("active", we["btn_delete_hover"])])

        self.style.configure("OpenCsv.TButton", padding=10, relief="flat", background=we["btn_open"], foreground="white")
        self.style.map("OpenCsv.TButton", background=[("active", we["btn_open_hover"])])

        self.style.configure("Exit.TButton", padding=10, relief="flat", background=we["btn_exit"], foreground="white")
        self.style.map("Exit.TButton", background=[("active", we["btn_exit_hover"])])

        # Progressbar
        self.style.configure("TProgressbar", troughcolor=p["progress_trough"], background=p["accent"])

        # Picker hover highlight (Combobox, Radio, Check) maps
        self.style.map("TCombobox",
                       fieldbackground=[("active", p["picker_highlight_bg"])],
                       foreground=[("active", p["picker_highlight_fg"])])
        self.style.map("TRadiobutton",
                       background=[("active", p["picker_highlight_bg"])],
                       foreground=[("active", p["picker_highlight_fg"])])
        self.style.map("TCheckbutton",
                       background=[("active", p["picker_highlight_bg"])],
                       foreground=[("active", p["picker_highlight_fg"])])

        # If text widget already exists, update its colors
        if hasattr(self, "log") and isinstance(self.log, Text):
            self.log.config(bg=p["text_bg"], fg=p["text_fg"], insertbackground=p["text_ins"])

        # Redraw gradient if exists
        if hasattr(self, "grad"):
            self._draw_gradient(self.grad, p["gradient_from"], p["gradient_to"])

    def _apply_picker_highlights(self, mode: str):
        """Ensure popdown list highlights for Combobox flip to opposite theme colors."""
        p = self.palettes[mode]
        try:
            self.root.option_add("*TCombobox*Listbox*selectBackground", p["picker_highlight_bg"])
            self.root.option_add("*TCombobox*Listbox*selectForeground", p["picker_highlight_fg"])
        except Exception:
            pass

    # ---------- UI composition ----------
    def _build_header(self):
        header = ttk.Frame(self.root, style="Panel.TFrame")
        header.pack(fill=X, side="top")

        # Gradient canvas
        self.grad = Canvas(header, height=86, highlightthickness=0, bd=0)
        self.grad.pack(fill=X, side="top")
        p = self.palettes[self.theme_var.get()]
        self._draw_gradient(self.grad, p["gradient_from"], p["gradient_to"])

        # Title and theme controls
        title_row = ttk.Frame(header, style="Panel.TFrame")
        title_row.pack(fill=X, padx=16, pady=(8, 12))

        left_box = ttk.Frame(title_row, style="Panel.TFrame")
        left_box.pack(side=LEFT, fill=X, expand=True)
        ttk.Label(left_box, text="Ragilmalik's Python GUI Mover", font=self.font_title).pack(anchor="w")
        ttk.Label(left_box, text="Clean, safe moves with metadata checks — no subfolders",
                  style="Muted.TLabel", font=self.font_subtle).pack(anchor="w", pady=(2, 0))

        right_box = ttk.Frame(title_row, style="Panel.TFrame")
        right_box.pack(side=RIGHT)

        ttk.Label(right_box, text="Theme:", style="Muted.TLabel", font=self.font_ui).pack(side=LEFT, padx=(0,6))
        self.theme_combo = ttk.Combobox(right_box, state="readonly", values=["Dark","Light"], width=8)
        self.theme_combo.set(self.theme_var.get())
        self.theme_combo.pack(side=LEFT)
        self.theme_combo.bind("<<ComboboxSelected>>", self._on_theme_change)

    def _build_controls(self):
        card = ttk.Frame(self.root, style="Card.TFrame")
        card.pack(fill=X, padx=16, pady=12)

        # Source
        row1 = ttk.Frame(card, style="Card.TFrame")
        row1.pack(fill=X, padx=12, pady=(12, 6))
        ttk.Label(row1, text="Source Folder:", font=self.font_ui).pack(side=LEFT)
        self.src_entry = ttk.Entry(row1, textvariable=self.src_var, width=80)
        self.src_entry.pack(side=LEFT, padx=8, fill=X, expand=True)
        ttk.Button(row1, text="Browse…", command=self.browse_src).pack(side=LEFT)

        # Destination
        row2 = ttk.Frame(card, style="Card.TFrame")
        row2.pack(fill=X, padx=12, pady=6)
        ttk.Label(row2, text="Destination Folder:", font=self.font_ui).pack(side=LEFT)
        self.dst_entry = ttk.Entry(row2, textvariable=self.dst_var, width=80)
        self.dst_entry.pack(side=LEFT, padx=8, fill=X, expand=True)
        ttk.Button(row2, text="Browse…", command=self.browse_dst).pack(side=LEFT)

        # Log location selector (xlsx)
        row3 = ttk.Frame(card, style="Card.TFrame")
        row3.pack(fill=X, padx=12, pady=(6, 0))

        ttk.Label(row3, text="Log File Location (.xlsx):", font=self.font_ui).pack(side=LEFT, padx=(0,10))
        self.csv_radio_dest = ttk.Radiobutton(row3, text="Destination Folder", value="dest", variable=self.csv_loc_var, command=self._csv_loc_changed)
        self.csv_radio_src  = ttk.Radiobutton(row3, text="Source Folder", value="source", variable=self.csv_loc_var, command=self._csv_loc_changed)
        self.csv_radio_cus  = ttk.Radiobutton(row3, text="Custom Folder", value="custom", variable=self.csv_loc_var, command=self._csv_loc_changed)
        self.csv_radio_dest.pack(side=LEFT, padx=(0,8))
        self.csv_radio_src.pack(side=LEFT, padx=(0,8))
        self.csv_radio_cus.pack(side=LEFT, padx=(0,8))

        row3b = ttk.Frame(card, style="Card.TFrame")
        row3b.pack(fill=X, padx=12, pady=(6, 12))
        self.custom_dir_entry = ttk.Entry(row3b, textvariable=self.csv_custom_dir, width=70, state="disabled")
        self.custom_dir_entry.pack(side=LEFT, padx=(0,8), fill=X, expand=True)
        self.custom_dir_btn = ttk.Button(row3b, text="Choose Folder…", command=self.choose_custom_dir, state="disabled")
        self.custom_dir_btn.pack(side=LEFT)

        # Options
        row4 = ttk.Frame(card, style="Card.TFrame")
        row4.pack(fill=X, padx=12, pady=(0, 12))
        self.dry_check = ttk.Checkbutton(row4, text="Simulation Only", variable=self.dry_run_var)
        self.dry_check.pack(side=LEFT)

        # Buttons row
        buttons = ttk.Frame(self.root, style="Card.TFrame")
        buttons.pack(fill=X, padx=16, pady=(0, 12))
        self.run_btn = ttk.Button(buttons, text="Run", style="Run.TButton", command=self.run)
        self.run_btn.pack(side=LEFT, padx=(12, 6))

        ttk.Button(buttons, text="Clear Log Screen", style="Clear.TButton", command=self.clear_log).pack(side=LEFT, padx=6)
        ttk.Button(buttons, text="Clear Log & Delete Last Log File", style="DeleteCsv.TButton", command=self.clear_log_and_delete_last_log).pack(side=LEFT, padx=6)
        ttk.Button(buttons, text="Open Last Saved Log File", style="OpenCsv.TButton", command=self.open_last_log).pack(side=LEFT, padx=6)

        ttk.Button(buttons, text="Exit", style="Exit.TButton", command=self.root.quit).pack(side=RIGHT, padx=12)

    def _build_progress(self):
        prog = ttk.Frame(self.root, style="Card.TFrame")
        prog.pack(fill=X, padx=16, pady=6)
        self.progress = ttk.Progressbar(prog, mode="determinate")
        self.progress.pack(fill=X, padx=12, pady=(12, 6))
        self.status_label = ttk.Label(prog, textvariable=self.status_var, style="Muted.TLabel", font=self.font_ui)
        self.status_label.pack(anchor="w", padx=12, pady=(0, 12))

    def _build_log(self):
        logframe_outer = ttk.Frame(self.root, style="Card.TFrame")
        logframe_outer.pack(fill=BOTH, expand=True, padx=16, pady=(0, 16))

        logframe = ttk.Frame(logframe_outer, style="Card.TFrame")
        logframe.pack(fill=BOTH, expand=True, padx=12, pady=12)

        p = self.palettes[self.theme_var.get()]
        self.log = Text(logframe, wrap="none", height=16, state=DISABLED,
            bg=p["text_bg"], fg=p["text_fg"], insertbackground=p["text_ins"], relief="flat")
        self.log.pack(side=LEFT, fill=BOTH, expand=True)
        vsb = ttk.Scrollbar(logframe, orient="vertical", command=self.log.yview)
        vsb.pack(side=RIGHT, fill=Y)
        self.log.configure(yscrollcommand=vsb.set)

    # ---------- Gradient header ----------
    def _draw_gradient(self, canvas, color1, color2):
        """Draw a left→right minimal split gradient in the given canvas."""
        canvas.update_idletasks()
        w = canvas.winfo_width() or self.root.winfo_width() or 960
        h = canvas.winfo_height() or 86
        canvas.delete("grad")
        canvas.create_rectangle(0, 0, w//2, h, fill=color1, outline="", tags="grad")
        canvas.create_rectangle(w//2, 0, w, h, fill=color2, outline="", tags="grad")
        canvas.bind("<Configure>", lambda e: self._draw_gradient(canvas, color1, color2))

    # ---------- Picker highlight handling ----------
    def _on_theme_change(self, _evt=None):
        self.theme_var.set(self.theme_combo.get())
        self._apply_theme(self.theme_var.get())
        self._apply_picker_highlights(self.theme_var.get())

    def _csv_loc_changed(self):
        mode = self.csv_loc_var.get()
        if mode == "custom":
            self.custom_dir_entry.config(state="normal")
            self.custom_dir_btn.config(state="normal")
        else:
            self.custom_dir_entry.config(state="disabled")
            self.custom_dir_btn.config(state="disabled")

    def choose_custom_dir(self):
        path = filedialog.askdirectory(title="Choose Custom Log Folder")
        if path:
            self.csv_custom_dir.set(path)

    # ---------- Helpers ----------
    def browse_src(self):
        path = filedialog.askdirectory(title="Select Source Folder")
        if path:
            self.src_var.set(path)

    def browse_dst(self):
        path = filedialog.askdirectory(title="Select Destination Folder")
        if path:
            self.dst_var.set(path)

    def clear_log(self):
        self.log.config(state=NORMAL)
        self.log.delete("1.0", END)
        self.log.config(state=DISABLED)

    def clear_log_and_delete_last_log(self):
        # Clear on-screen log
        self.clear_log()

        # Delete last XLSX if present
        if not self._last_log_path:
            messagebox.showinfo("Delete Last Log", "No log file has been created yet this session.")
            return

        try:
            if os.path.exists(self._last_log_path):
                os.remove(self._last_log_path)
                deleted = self._last_log_path
                self._last_log_path = None
                messagebox.showinfo("Delete Last Log", f"Deleted: {deleted}")
            else:
                messagebox.showinfo("Delete Last Log", "Last log file path does not exist anymore.")
                self._last_log_path = None
        except Exception as e:
            messagebox.showerror("Delete Last Log", f"Could not delete log file:\n{e}")

    def open_last_log(self):
        if not self._last_log_path or not os.path.exists(self._last_log_path):
            messagebox.showinfo("Open Last Log", "No saved log file found for this session.")
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(self._last_log_path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                os.system(f'open "{self._last_log_path}"')
            else:
                try:
                    os.system(f'xdg-open "{self._last_log_path}"')
                except Exception:
                    webbrowser.open(f"file://{self._last_log_path}")
        except Exception as e:
            messagebox.showerror("Open Last Log", f"Could not open log file:\n{e}")

    def append_log(self, text):
        self.log.config(state=NORMAL)
        self.log.insert(END, text + "\n")
        self.log.see(END)
        self.log.config(state=DISABLED)

    def set_running(self, running: bool):
        self.run_btn.config(state=DISABLED if running else NORMAL)

    # ---------- XLSX helpers ----------
    def _resolve_log_dir(self, src_dir, dst_dir):
        mode = self.csv_loc_var.get()
        if mode == "source":
            return src_dir
        elif mode == "dest":
            return dst_dir
        else:  # custom
            chosen = self.csv_custom_dir.get().strip()
            if not chosen:
                messagebox.showinfo("Log Folder", "Please choose a custom folder to save the Excel log.")
                self.choose_custom_dir()
                chosen = self.csv_custom_dir.get().strip()
            if not chosen:
                raise ValueError("Custom log folder not selected.")
            if not os.path.isdir(chosen):
                raise ValueError("Custom log folder is not a valid directory.")
            return chosen

    def _open_xlsx_log(self, out_dir):
        if Workbook is None:
            raise RuntimeError("openpyxl is required to write .xlsx logs. Please install it: pip install openpyxl")
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"SmartFileMover-log-{ts}.xlsx"
        path = os.path.join(out_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Log"

        header = [
            "Timestamp", "Action", "Source Folder", "Destination Folder",
            "Filename", "New Filename", "File Creation Time",
            "Size", "Note"
        ]
        ws.append(header)

        # Header style (bold)
        bold = Font(bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold
            cell.alignment = Alignment(vertical="center")

        self._last_log_path = path

        # Remember column indexes for formatting
        self.COL_TIMESTAMP = 1
        self.COL_FILE_CTIME = 7
        self.COL_SIZE = 8

        return wb, ws, path

    def _autosize_columns(self, ws):
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 80)

    def _write_log_row(self, ws, src_dir, dst_dir, action, filename, new_filename, ctime, size_bytes, note=""):
        """
        Write one log row and apply formatting:
        - Timestamp (datetime object with Excel number format DD/MM/YYYY HH:MM:SS)
        - File Creation Time (datetime with format DD/MM/YYYY HH:MM:SS)
        - Size as text 'NNNKB' (no decimals)
        """
        ts_now_dt = datetime.now()

        # Size text in KB (integer, with 'KB' suffix)
        if isinstance(size_bytes, (int, float)) and size_bytes >= 0:
            size_kb_int = int(round(size_bytes / 1024.0))
            size_text = f"{size_kb_int}KB"
        else:
            size_text = ""

        # Convert ctime -> datetime
        ctime_dt = datetime.fromtimestamp(ctime) if isinstance(ctime, (int, float)) and ctime >= 0 else None

        row_data = [
            ts_now_dt,             # Timestamp (datetime)
            action,
            src_dir,
            dst_dir,
            filename or "",
            new_filename or "",
            ctime_dt if ctime_dt else "",  # File Creation Time (datetime or empty)
            size_text,              # Size as 'NNNKB' text
            note or ""
        ]
        ws.append(row_data)

        # Apply number formats to the just-appended row
        r = ws.max_row
        ts_cell = ws.cell(row=r, column=self.COL_TIMESTAMP)
        ts_cell.number_format = "DD/MM/YYYY HH:MM:SS"

        ctime_cell = ws.cell(row=r, column=self.COL_FILE_CTIME)
        if ctime_dt:
            ctime_cell.number_format = "DD/MM/YYYY HH:MM:SS"

    # ---------- Core flow ----------
    def run(self):
        src = self.src_var.get().strip()
        dst = self.dst_var.get().strip()

        if not src or not dst:
            messagebox.showerror("Error", "Please select both source and destination folders.")
            return
        if not os.path.isdir(src):
            messagebox.showerror("Error", "Source folder does not exist or is not a directory.")
            return
        if not os.path.isdir(dst):
            messagebox.showerror("Error", "Destination folder does not exist or is not a directory.")
            return
        if os.path.abspath(src) == os.path.abspath(dst):
            messagebox.showerror("Error", "Source and destination folders must be different.")
            return

        # Resolve log folder *before* starting
        try:
            log_dir = self._resolve_log_dir(src, dst)
        except Exception as e:
            messagebox.showerror("Log Folder", f"Cannot proceed: {e}")
            return

        self.set_running(True)
        mode = "DRY RUN" if self.dry_run_var.get() else "LIVE RUN"
        self.status_var.set(f"Preparing ({mode})…")

        self.append_log(f"Starting {mode.lower()} from:\n  {src}\n→ {dst}\n(no subfolders)")
        self.append_log(f"Excel log will be saved to: {log_dir}")

        # Start worker
        self._worker_thread = threading.Thread(
            target=self._worker, args=(src, dst, log_dir, self.dry_run_var.get()), daemon=True
        )
        self._worker_thread.start()

    def _worker(self, src, dst, log_dir, dry_run):
        wb = None
        ws = None
        try:
            try:
                wb, ws, log_path = self._open_xlsx_log(log_dir)
            except Exception as e:
                self.append_log(f"WARNING: Could not create Excel log. {e}")
                wb = ws = None

            files = [f for f in os.listdir(src) if os.path.isfile(os.path.join(src, f))]
            total = len(files)
            moved = 0
            skipped = 0
            errors = 0

            self.progress["value"] = 0
            self.progress["maximum"] = total if total > 0 else 1

            if total == 0:
                self.append_log("No files found in source (top-level only). Nothing to do.")
                self.status_var.set("Done (no files).")
                if ws:
                    self._write_log_row(ws, src, dst, "INFO", "", "", -1, -1, "No files to process")
                    self._autosize_columns(ws)
                    wb.save(self._last_log_path)
                return

            for idx, name in enumerate(files, start=1):
                src_path = os.path.join(src, name)
                planned_dst_same = os.path.join(dst, name)

                try:
                    s_name, s_ext, s_size, s_mtime, s_ctime = file_meta(src_path)

                    if os.path.exists(planned_dst_same):
                        if files_identical(src_path, planned_dst_same):
                            msg = (f"SKIP: {name} (identical) "
                                   f"[ext={s_ext}, size={s_size}, mtime={human_time(s_mtime)}]")
                            self.append_log(msg)
                            skipped += 1
                            if ws:
                                self._write_log_row(ws, src, dst, "SKIP", name, "", s_ctime, s_size, "Identical metadata")
                        else:
                            new_name = next_available_name(dst, name)
                            final_dst = os.path.join(dst, new_name)
                            if dry_run:
                                self.append_log(f"DRYRUN: would move (renamed) {name} → {new_name}")
                                moved += 1
                                if ws:
                                    self._write_log_row(ws, src, dst, "DRYRUN_MOVED_RENAMED", name, new_name, s_ctime, s_size, "Different metadata; rename required")
                            else:
                                shutil.move(src_path, final_dst)
                                self.append_log(f"MOVED (renamed): {name} → {new_name}")
                                moved += 1
                                if ws:
                                    self._write_log_row(ws, src, dst, "MOVED_RENAMED", name, new_name, s_ctime, s_size, "Different metadata; renamed")
                    else:
                        if dry_run:
                            self.append_log(f"DRYRUN: would move {name}")
                            moved += 1
                            if ws:
                                self._write_log_row(ws, src, dst, "DRYRUN_MOVED", name, "", s_ctime, s_size, "")
                        else:
                            shutil.move(src_path, planned_dst_same)
                            self.append_log(f"MOVED: {name}")
                            moved += 1
                            if ws:
                                self._write_log_row(ws, src, dst, "MOVED", name, "", s_ctime, s_size, "")
                except Exception as e:
                    errors += 1
                    err_msg = f"ERROR moving {name}: {e}"
                    self.append_log(err_msg)
                    if ws:
                        try:
                            _, _, s_size, _, s_ctime = file_meta(src_path)
                        except Exception:
                            s_size, s_ctime = -1, -1
                        self._write_log_row(ws, src, dst, "ERROR", name, "", s_ctime, s_size, str(e))

                self.progress["value"] = idx
                run_mode = "DRY RUN" if dry_run else "LIVE RUN"
                self.status_var.set(f"{run_mode}: Processed {idx}/{total}…")

            # Empty row then SUMMARY row (bold)
            if ws:
                ws.append([""] * 9)
                summary_note = (
                    f"Summary — planned_or_moved={moved}, skipped={skipped}, "
                    f"errors={errors}, total={total}; mode={'Simulation Only' if dry_run else 'Live Run'}"
                )
                self._write_log_row(ws, src, dst, "SUMMARY", "", "", -1, -1, summary_note)

                # Bold the entire SUMMARY row
                summary_row_idx = ws.max_row
                for col in range(1, 10):  # 9 columns
                    cell = ws.cell(row=summary_row_idx, column=col)
                    cell.font = Font(bold=True)

                # Autosize & save
                self._autosize_columns(ws)
                wb.save(self._last_log_path)
                self.append_log(f"Log saved: {self._last_log_path}")

            self.append_log("-" * 70)
            self.append_log(f"Summary: planned_or_moved={moved}, skipped={skipped}, errors={errors}, total={total}")
            self.status_var.set("Done.")
        finally:
            self.set_running(False)

# ----------------------------
# Entrypoint
# ----------------------------
def main():
    # Enable high-DPI awareness on Windows so the UI isn’t blurry in the exe
    if sys.platform.startswith("win"):
        try:
            import ctypes
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

    root = Tk()
    app = MoveApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
